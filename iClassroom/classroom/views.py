from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from .models import User
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from .forms import UserProfileForm
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.contrib.auth.tokens import default_token_generator
from django.urls import reverse
from django.utils.safestring import mark_safe  # Add this import
import logging
import pytz
from pytz import timezone as pytz_timezone
from django.utils import timezone
from django.db import transaction
import json
from datetime import datetime, timedelta
from django.db import models
from django.db.models import Count, Avg, Sum, F,FloatField,Q
from .models import Class, Notification, ClassSession, Enrollment, Attendance, Participation, ParticipationCategory, QRCode
from .forms import ClassForm ,SessionForm
from django.views.generic import ListView, CreateView, UpdateView
from django.http import JsonResponse
from django.urls import reverse_lazy
from collections import defaultdict 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
import qrcode
logger = logging.getLogger(__name__)

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        # Check if user exists and is inactive
        try:
            user_obj = User.objects.get(email=email)
            if not user_obj.is_active:
                # Use mark_safe to allow HTML in the message
                error_message = mark_safe(
                    'Your account is not yet verified. Please verify your email address by clicking the link sent to your Gmail inbox. '
                    'If you did not receive the email, <a href="{0}?email={1}">click here to resend verification</a>.'
                    .format(reverse('resend_verification'), email)
                )
                messages.error(request, error_message)
                return render(request, 'classroom/login.html')
        except User.DoesNotExist:
            pass  # Continue to normal authentication

        user = authenticate(request, email=email, password=password)

        if user is not None:
            login(request, user)
            if user.user_type == 'teacher':
                return redirect('teacher_dashboard')
            elif user.user_type == 'student':
                return redirect('student_dashboard')
            else:
                return redirect('dashboard')
        else:
            messages.error(request, 'Invalid email or password.')

    return render(request, 'classroom/login.html')


def send_verification_email(request, user):
    """Send verification email to user"""
    try:
        current_site = get_current_site(request)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        token = default_token_generator.make_token(user)
        verification_url = request.build_absolute_uri(
            reverse('verify_email', kwargs={'uidb64': uid, 'token': token})
        )
        subject = 'Verify your iClassroom account'
        html_message = render_to_string('classroom/emails/verification_email.html', {
            'user': user,
            'verification_url': verification_url,
            'site_name': current_site.name,
        })
        plain_message = strip_tags(html_message)
        logger.debug(f"Sending email to {user.email} with verification URL: {verification_url}")
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            html_message=html_message,
            fail_silently=False,
        )
        logger.info(f"Verification email sent to {user.email}")
        return True
    except Exception as e:
        logger.error(f"Failed to send verification email to {user.email}: {str(e)}", exc_info=True)
        return False


def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        user_type = request.POST.get('user_type')
        
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already registered.")
            return redirect('register')

        # Create inactive user
        user = User.objects.create_user(
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            user_type=user_type,
            is_active=False  # User must verify email first
        )
        
        # Send verification email
        if send_verification_email(request, user):
            messages.success(request, f"Registration successful! Please check your email ({email}) for a verification link.")
        else:
            messages.error(request, "Registration successful but failed to send verification email. Please contact support.")
        
        return redirect('registration_complete')
    
    return render(request, 'classroom/register.html')


def registration_complete(request):
    """Show registration complete page"""
    return render(request, 'classroom/registration_complete.html')


def verify_email(request, uidb64, token):
    """Verify email address"""
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
        
        if default_token_generator.check_token(user, token):
            if not user.is_active:
                user.is_active = True
                user.save()
                messages.success(request, 'Email verified successfully! You can now log in.')
            else:
                messages.info(request, 'Email already verified.')
            return redirect('login')
        else:
            messages.error(request, 'Invalid verification link.')
            return redirect('resend_verification')
            
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        messages.error(request, 'Invalid verification link.')
        return redirect('resend_verification')


def resend_verification(request):
    """Resend verification email"""
    email = request.GET.get('email', '')  # Prefill if present
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email, is_active=False)
            if send_verification_email(request, user):
                messages.success(request, f'Verification email resent to {email}')
            else:
                messages.error(request, 'Failed to send verification email. Please try again.')
        except User.DoesNotExist:
            messages.error(request, 'No inactive account found with this email address.')
        return redirect('resend_verification')
    return render(request, 'classroom/resend_verification.html', {'email': email})


@login_required
def dashboard_view(request):
    if request.user.user_type == 'teacher':
        return redirect('teacher_dashboard')
    elif request.user.user_type == 'student':
        return redirect('student_dashboard')
    else:
        return render(request, 'classroom/dashboard.html')

@login_required
def teacher_dashboard(request):
    """Teacher dashboard view with create class modal"""
    manila_tz = pytz_timezone('Asia/Manila')
    now = timezone.now().astimezone(manila_tz)
    current_date = now.date()
    current_time = now.time()

    if request.method == 'POST':
        # Handle AJAX form submission for creating class
        try:
            form = ClassForm(request.POST)
            if form.is_valid():
                new_class = form.save(commit=False)
                new_class.teacher = request.user
                new_class.save()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Class created successfully!',
                    'class_name': f"{new_class.course_name} - {new_class.section}",
                    'class_id': new_class.id
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Please check the form data and try again.',
                    'form_errors': form.errors
                })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })

    # Get all active classes taught by this teacher
    teacher_classes = Class.objects.filter(
        teacher=request.user,
        is_active=True
    )
    # Annotate each class with additional data
    classes_with_stats = []
    for class_obj in teacher_classes:
        # Get enrollment count
        enrollment_count = Enrollment.objects.filter(
            class_instance=class_obj,
            status='active'
        ).count()

        # Calculate attendance percentage (PH timezone aware)
        completed_sessions = ClassSession.objects.filter(
            class_instance=class_obj,
            status='completed'
        )

        total_possible_attendances = 0
        present_attendances = 0
        
        for session in completed_sessions:
            # Convert session times to Manila timezone for accurate comparison
            session_date_manila = session.session_date
            session_start_manila = session.start_time
            session_end_manila = session.end_time
            
            session_attendances = Attendance.objects.filter(
                session=session
            ).count()
            total_possible_attendances += enrollment_count
            present_attendances += session_attendances

        attendance_percentage = 0
        if total_possible_attendances > 0:
            attendance_percentage = (present_attendances / total_possible_attendances) * 100

        # Calculate participation average
        participation_avg = Participation.objects.filter(
            session__class_instance=class_obj
        ).aggregate(
            avg=Avg('points'),
            count=Count('id')
        )
        participation_average = participation_avg['avg'] or 0
        participation_count = participation_avg['count'] or 0

        # Get next session (PH timezone aware)
        next_session = ClassSession.objects.filter(
            class_instance=class_obj,
            session_date__gte=current_date,
            status='scheduled'  # Only get scheduled sessions
        ).exclude(
            session_date=current_date,
            start_time__lte=current_time,
            end_time__gte=current_time
        ).order_by('session_date', 'start_time').first()

        # Get current session (PH timezone aware)
        current_session = ClassSession.objects.filter(
            class_instance=class_obj,
            session_date=current_date,
            start_time__lte=current_time,
            end_time__gte=current_time,
            status='in_progress'
        ).first()

        # If no current session found, check if there should be one
        if not current_session:
            potential_session = ClassSession.objects.filter(
                class_instance=class_obj,
                session_date=current_date,
                start_time__lte=current_time,
                end_time__gte=current_time
            ).first()
            if potential_session:
                potential_session.status = 'in_progress'
                potential_session.save()
                current_session = potential_session

        classes_with_stats.append({
            'object': class_obj,
            'enrollment_count': enrollment_count,
            'attendance_percentage': round(attendance_percentage, 1),
            'participation_average': round(participation_average, 1),
            'participation_count': participation_count,
            'next_session': next_session,
            'current_session': current_session
        })
    # Recent activities (notifications)
    recent_activities = Notification.objects.filter(
        user=request.user
    ).order_by('-created_at')[:5]

    # Statistics
    total_students = sum(c['enrollment_count'] for c in classes_with_stats)

    avg_attendance = 0
    if classes_with_stats:
        avg_attendance = sum(c['attendance_percentage'] for c in classes_with_stats) / len(classes_with_stats)

    active_classes = len(classes_with_stats)

    # Calculate session hours manually for this month
    sessions_this_month = ClassSession.objects.filter(
        class_instance__teacher=request.user,
        session_date__month=timezone.now().month
    )

    total_duration = timedelta()
    for session in sessions_this_month:
        start = datetime.combine(session.session_date, session.start_time)
        end = datetime.combine(session.session_date, session.end_time)
        total_duration += (end - start)

    session_hours = total_duration.total_seconds() / 3600  # Convert to hours

    form = ClassForm()

    context = {
        'classes_with_stats': classes_with_stats,
        'recent_activities': recent_activities,
        'total_students': total_students,
        'average_attendance': round(avg_attendance, 1),
        'active_classes': active_classes,
        'session_hours': round(session_hours, 1),
        'form': form,
    }

    return render(request, 'classroom/teacher/teacher_dashboard.html', context)

@login_required
def class_detail(request, class_id):
    # Get the class or return 404 if not found or not owned by this teacher
    class_obj = get_object_or_404(Class, id=class_id, teacher=request.user)
    
    # Get all active enrollments for this class
    enrollments = Enrollment.objects.filter(
        class_instance=class_obj,
        status='active'
    ).select_related('student')
    
    # Get upcoming sessions (next 5)
    upcoming_sessions = ClassSession.objects.filter(
        class_instance=class_obj,
        session_date__gte=timezone.now().date()
    ).order_by('session_date', 'start_time')[:5]
    
    # Get recent sessions (last 5 completed)
    recent_sessions = ClassSession.objects.filter(
        class_instance=class_obj,
        status='completed'
    ).order_by('-session_date', '-start_time')[:5]
    
    # Calculate attendance stats per student
    attendance_stats = []
    for enrollment in enrollments:
        total_sessions = ClassSession.objects.filter(
            class_instance=class_obj,
            status='completed'
        ).count()
        
        present_count = Attendance.objects.filter(
            student=enrollment.student,
            session__class_instance=class_obj,
            status='present'
        ).count()
        
        percentage = 0
        dashoffset = 163.36  # full circle default
        if total_sessions > 0:
            percentage = round((present_count / total_sessions) * 100, 1)
            dashoffset = round(163.36 - (percentage * 1.6336), 2)

        attendance_stats.append({
            'enrollment': enrollment,
            'present_count': present_count,
            'total_sessions': total_sessions,
            'percentage': percentage,
            'dashoffset': dashoffset
        })
    
    # Calculate participation stats per student with category breakdown
    participation_stats = []
    for enrollment in enrollments:
        # Get total points per category
        categories = ParticipationCategory.objects.filter(
            class_instance=class_obj
        )
        
        category_points = {}
        for category in categories:
            points = Participation.objects.filter(
                student=enrollment.student,
                session__class_instance=class_obj,
                category=category
            ).aggregate(total=Sum('points'))['total'] or 0
            category_points[category.name.lower()] = points
        
        # Get total points across all categories
        total_points = sum(category_points.values())
        
        participation_stats.append({
            'enrollment': enrollment,
            'total_points': total_points,
            'category_points': category_points
        })
    
    # Calculate averages for the class
    total_students = enrollments.count()
    avg_attendance = 0
    avg_participation = 0
    
    if total_students > 0:
        total_attendance = sum(stat['percentage'] for stat in attendance_stats)
        avg_attendance = round(total_attendance / total_students, 1)
        
        total_participation = sum(stat['total_points'] for stat in participation_stats)
        avg_participation = round(total_participation / total_students, 1)
    
    # Combine both attendance and participation per student using zip
    student_stats = list(zip(attendance_stats, participation_stats))

    # Get recent activities/notifications related to this class
    recent_activities = Notification.objects.filter(
        Q(user=request.user) & 
        (Q(related_class=class_obj) | Q(notification_type='class'))
    ).order_by('-created_at')[:10]
    # Analytics Data
    attendance_trend = []
    participation_trend = []
    for i in range(4, -1, -1):
        date = timezone.now() - timedelta(weeks=i)
        week_start = date - timedelta(days=date.weekday())
        week_end = week_start + timedelta(days=6)
        
        # Weekly attendance
        weekly_sessions = ClassSession.objects.filter(
            class_instance=class_obj,
            session_date__range=[week_start, week_end]
        )
        present_count = Attendance.objects.filter(
            session__in=weekly_sessions,
            status='present'
        ).count()
        attendance_trend.append({
            'week': week_start.strftime("%b %d"),
            'present': present_count,
            'total': weekly_sessions.count()
        })
        
        # Weekly participation
        weekly_points = Participation.objects.filter(
            session__in=weekly_sessions
        ).aggregate(total=Sum('points'))['total'] or 0
        participation_trend.append({
            'week': week_start.strftime("%b %d"),
            'points': weekly_points
        })
    
    # Participation by category
    category_distribution = Participation.objects.filter(
        session__class_instance=class_obj
    ).values('category__name').annotate(
        total_points=Sum('points')
    ).order_by('-total_points')
    
    
    context = {
        'class': class_obj,
        'enrollments': enrollments,
        'upcoming_sessions': upcoming_sessions,
        'recent_sessions': recent_sessions,
        'student_stats': student_stats,
        'recent_activities': recent_activities,
        'avg_attendance': avg_attendance,
        'avg_participation': avg_participation,
        'total_students': total_students,
        'attendance_trend': attendance_trend,
        'participation_trend': participation_trend,
        'category_distribution': category_distribution,
        'report_types': ['Attendance', 'Participation', 'Combined', 'Student Performance']
    }
    
    return render(request, 'classroom/class_detail.html', context)
@login_required
def student_dashboard(request):
    student = request.user
    
    # Set Manila timezone
    manila_tz = pytz_timezone('Asia/Manila')
    now = timezone.now().astimezone(manila_tz)
    current_date = now.date()
    current_time = now.time()
    
    # Get all active enrollments for this student
    enrollments = Enrollment.objects.filter(
        student=student,
        status='active'
    ).select_related('class_instance')
    
    # Calculate overall statistics
    total_classes = enrollments.count()
    
    # Calculate classes today (PH time)
    classes_today = ClassSession.objects.filter(
        class_instance__in=[e.class_instance for e in enrollments],
        session_date=current_date
    ).count()
    
    # Calculate attendance statistics
    attendance_stats = []
    overall_attendance_percentage = 0
    overall_participation_score = 0
    
    if enrollments.exists():
        for enrollment in enrollments:
            class_obj = enrollment.class_instance
            
            # Get completed sessions for this class
            completed_sessions = ClassSession.objects.filter(
                class_instance=class_obj,
                status='completed'
            )
            
            # Get attendance records for this student
            attendance_records = Attendance.objects.filter(
                session__in=completed_sessions,
                student=student
            )
            
            # Calculate attendance percentage
            total_sessions = completed_sessions.count()
            present_count = attendance_records.filter(status='present').count()
            attendance_percentage = (present_count / total_sessions * 100) if total_sessions > 0 else 0
            
            # Calculate participation score
            participation_records = Participation.objects.filter(
                session__class_instance=class_obj,
                student=student
            )
            participation_points = participation_records.aggregate(total=Sum('points'))['total'] or 0
            participation_avg = participation_records.aggregate(avg=Avg('points'))['avg'] or 0
            
            # Get next session (PH time)
            next_session = ClassSession.objects.filter(
                class_instance=class_obj,
                session_date__gte=current_date,
                status='scheduled'  # Only get scheduled sessions
            ).exclude(
                session_date=current_date,
                start_time__lte=current_time,
                end_time__gte=current_time
            ).order_by('session_date', 'start_time').first()
            
            # Check for current session (PH time)
            current_session = ClassSession.objects.filter(
                class_instance=class_obj,
                session_date=current_date,
                start_time__lte=current_time,
                end_time__gte=current_time,
                status='in_progress'
            ).first()
            
            # If no current session found, check if there should be one
            if not current_session:
                potential_session = ClassSession.objects.filter(
                    class_instance=class_obj,
                    session_date=current_date,
                    start_time__lte=current_time,
                    end_time__gte=current_time
                ).first()
                if potential_session:
                    potential_session.status = 'in_progress'
                    potential_session.save()
                    current_session = potential_session
            
            attendance_stats.append({
                'class': class_obj,
                'attendance_percentage': round(attendance_percentage, 1),
                'participation_score': round(participation_avg, 1),
                'next_session': next_session,
                'current_session': current_session,
                'total_sessions': total_sessions,
                'present_count': present_count,
                'participation_points': participation_points
            })
        
        # Calculate overall attendance percentage
        total_present = sum(s['present_count'] for s in attendance_stats)
        total_sessions = sum(s['total_sessions'] for s in attendance_stats)
        overall_attendance_percentage = (total_present / total_sessions * 100) if total_sessions > 0 else 0
        
        # Calculate overall participation score
        overall_participation_score = sum(s['participation_score'] for s in attendance_stats) / len(attendance_stats) if attendance_stats else 0
    
    # Get recent activities with PH time
    recent_activities = []
    attendance_activities = Attendance.objects.filter(
        student=student
    ).select_related('session', 'session__class_instance').order_by('-marked_at')[:5]
    
    participation_activities = Participation.objects.filter(
        student=student
    ).select_related('session', 'session__class_instance', 'category').order_by('-created_at')[:5]
    
    # Combine and sort activities with PH time
    for activity in attendance_activities:
        recent_activities.append({
            'type': 'attendance',
            'activity': activity,
            'time': activity.marked_at.astimezone(manila_tz)
        })
    
    for activity in participation_activities:
        recent_activities.append({
            'type': 'participation',
            'activity': activity,
            'time': activity.created_at.astimezone(manila_tz)
        })
    
    # Sort combined activities by time
    recent_activities.sort(key=lambda x: x['time'], reverse=True)
    recent_activities = recent_activities[:5]
    
    context = {
        'enrollments': enrollments,
        'attendance_stats': attendance_stats,
        'total_classes': total_classes,
        'classes_today': classes_today,
        'overall_attendance_percentage': round(overall_attendance_percentage, 1),
        'overall_participation_score': round(overall_participation_score, 1),
        'recent_activities': recent_activities,
        'current_ph_time': now.strftime("%b %d, %Y %I:%M %p"),  # For display in template
    }
    
    return render(request, 'classroom/student/student_dashboard.html', context)
@login_required
def enroll_in_class(request):
    if request.method == 'POST':
        class_id = request.POST.get('class_id')
        
        try:
            class_to_join = Class.objects.get(id=class_id, is_active=True)
            
            # Check if student is already enrolled
            if Enrollment.objects.filter(student=request.user, class_instance=class_to_join).exists():
                messages.warning(request, 'You are already enrolled in this class')
            else:
                # Create new enrollment
                Enrollment.objects.create(
                    student=request.user,
                    class_instance=class_to_join,
                    status='active'
                )
                messages.success(request, f'Successfully enrolled in {class_to_join.course_name}')
                
        except Class.DoesNotExist:
            messages.error(request, 'Invalid class ID. Please check with your teacher')
        except Exception as e:
            messages.error(request, 'An error occurred. Please try again')
            print(f"Enrollment error: {e}")
            
    return redirect('student_dashboard')
@login_required
def student_progress(request):
    student = request.user
    enrollments = Enrollment.objects.filter(student=student, status='active').select_related('class_instance')
    attendance_stats = []

    for enrollment in enrollments:
        class_obj = enrollment.class_instance

        # Attendance
        completed_sessions = ClassSession.objects.filter(class_instance=class_obj, status='completed')
        total_sessions = completed_sessions.count()
        attended_sessions = Attendance.objects.filter(
            session__in=completed_sessions,
            student=student,
            status='present'
        ).count()
        attendance_percentage = round((attended_sessions / total_sessions) * 100, 1) if total_sessions > 0 else 0

        # Participation
        participation_records = Participation.objects.filter(
            session__class_instance=class_obj,
            student=student
        )
        participation_points = participation_records.aggregate(total=Sum('points'))['total'] or 0
        # You can set max_participation_points as needed, e.g. 200 or calculate from class rules
        max_participation_points = 200
        participation_percentage = round((participation_points / max_participation_points) * 100, 1) if max_participation_points > 0 else 0

        # Grade (dummy logic, replace with your own)
        grade = "A" if attendance_percentage >= 90 and participation_percentage >= 90 else "A-" if attendance_percentage >= 80 else "B"

        attendance_stats.append({
            'class': class_obj,
            'attended_sessions': attended_sessions,
            'total_sessions': total_sessions,
            'attendance_percentage': attendance_percentage,
            'participation_points': participation_points,
            'max_participation_points': max_participation_points,
            'participation_percentage': participation_percentage,
            'grade': grade,
            'teacher_name': f"{class_obj.teacher.first_name} {class_obj.teacher.last_name}",
            'course_code': getattr(class_obj, 'course_code', ''),
        })

    # --- Add this block to fetch recent activities ---
    manila_tz = pytz_timezone('Asia/Manila')
    attendance_activities = Attendance.objects.filter(
        student=student
    ).select_related('session', 'session__class_instance').order_by('-marked_at')[:5]

    participation_activities = Participation.objects.filter(
        student=student
    ).select_related('session', 'session__class_instance', 'category').order_by('-created_at')[:5]

    recent_activities = []
    for activity in attendance_activities:
        recent_activities.append({
            'type': 'attendance',
            'activity': activity,
            'time': activity.marked_at.astimezone(manila_tz)
        })
    for activity in participation_activities:
        recent_activities.append({
            'type': 'participation',
            'activity': activity,
            'time': activity.created_at.astimezone(manila_tz)
        })
    recent_activities.sort(key=lambda x: x['time'], reverse=True)
    recent_activities = recent_activities[:5]
    # -------------------------------------------------

    context = {
        'attendance_stats': attendance_stats,
        'recent_activities': recent_activities,  # Add this line
    }
    return render(request, 'classroom/student/student_progress.html', context)
@login_required
def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def profile_view(request):
    """Display user profile page"""
    context = {
        'user': request.user,
        'user_type': getattr(request.user, 'user_type', 'student'),
    }

    return render(request, 'classroom/profile.html', context)

@login_required
def profile_edit(request):
    """Edit user profile"""
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            user = form.save(commit=False)

            # Save avatar separately if provided
            if 'avatar' in request.FILES:
                avatar = request.FILES['avatar']
                user.avatar_url = f"/media/avatars/{avatar.name}"
                with open(f"media/avatars/{avatar.name}", 'wb+') as destination:
                    for chunk in avatar.chunks():
                        destination.write(chunk)

            user.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('profile')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = UserProfileForm(instance=request.user)

    return render(request, 'classroom/profile_edit.html', {'form': form, 'user': request.user})

@login_required
@require_http_methods(["POST"])
def upload_avatar(request):
    """Handle avatar upload via AJAX"""
    try:
        if 'avatar' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No file provided'})

        avatar = request.FILES['avatar']

        if not avatar.content_type.startswith('image/'):
            return JsonResponse({'success': False, 'error': 'File must be an image'})

        if avatar.size > 5 * 1024 * 1024:
            return JsonResponse({'success': False, 'error': 'File size must be < 5MB'})

        # Save avatar file manually
        path = f"media/avatars/{avatar.name}"
        with open(path, 'wb+') as f:
            for chunk in avatar.chunks():
                f.write(chunk)

        # Save avatar URL to user
        request.user.avatar_url = f"/media/avatars/{avatar.name}"
        request.user.save()

        return JsonResponse({
            'success': True,
            'avatar_url': request.user.avatar_url,
            'message': 'Avatar updated successfully!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def remove_avatar(request):
    """Remove user avatar"""
    try:
        if request.user.avatar_url:
            # Optional: delete actual file from storage
            import os
            avatar_path = request.user.avatar_url.replace("/media/", "media/")
            if os.path.exists(avatar_path):
                os.remove(avatar_path)

            request.user.avatar_url = None
            request.user.save()

        return JsonResponse({
            'success': True,
            'message': 'Avatar removed successfully!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def change_password(request):
    """Change user password"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Password changed successfully!')
            return redirect('settings')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordChangeForm(request.user)
    
    context = {
        'form': form,
        'user': request.user,
    }
    
    return render(request, 'classroom/change_password.html', context)

@login_required
@require_http_methods(["POST"])
def delete_account(request):
    """Delete user account"""
    if request.method == 'POST':
        password = request.POST.get('password')
        if request.user.check_password(password):
            request.user.delete()
            messages.success(request, 'Account deleted successfully.')
            return redirect('login')
        else:
            messages.error(request, 'Incorrect password.')
            return redirect('settings')
    
    return redirect('settings')

def forgot_password(request):
    """Handle forgot password functionality"""
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            if not user.is_active:
                messages.error(request, 'You must verify your email before you can reset your password.')
                return redirect('forgot_password')
            current_site = get_current_site(request)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)
            reset_url = request.build_absolute_uri(
                reverse('reset_password', kwargs={'uidb64': uid, 'token': token})
            )
            subject = 'Reset your iClassroom password'
            html_message = render_to_string('classroom/emails/reset_password_email.html', {
                'user': user,
                'reset_url': reset_url,
                'site_name': current_site.name,
            })
            plain_message = strip_tags(html_message)
            send_mail(
                subject=subject,
                message=plain_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                html_message=html_message,
                fail_silently=False,
            )
            messages.success(request, f'Password reset link sent to {email}.')
        except User.DoesNotExist:
            messages.error(request, 'No account found with that email address.')
        return redirect('forgot_password')
    
    return render(request, 'classroom/forgot_password.html')

def reset_password(request, uidb64, token):
    """Reset user password"""
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)

        if not user.is_active:
            messages.error(request, 'You must verify your email before you can reset your password.')
            return redirect('login')

        if default_token_generator.check_token(user, token):
            if request.method == 'POST':
                new_password = request.POST.get('new_password')
                confirm_password = request.POST.get('confirm_password')

                if new_password == confirm_password:
                    user.set_password(new_password)
                    user.save()
                    messages.success(request, 'Password reset successfully! You can now log in.')
                    return redirect('login')
                else:
                    messages.error(request, 'Passwords do not match.')
            return render(request, 'classroom/reset_password.html', {'uidb64': uidb64, 'token': token})
        else:
            messages.error(request, 'Invalid password reset link.')
            return redirect('forgot_password')
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        messages.error(request, 'Invalid password reset link.')
        return redirect('forgot_password')

class SessionListView(ListView):
    model = ClassSession
    template_name = 'classroom/teacher/session_list.html'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # Auto-close expired sessions when listing
        for session in queryset.filter(status='in_progress'):
            session.auto_complete_if_expired()
        return queryset.filter(class_instance_id=self.kwargs['class_id']).order_by('-session_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['class_obj'] = self.class_obj = get_object_or_404(Class, id=self.kwargs['class_id'])
        context['now_ph'] = timezone.now().astimezone(pytz.timezone('Asia/Manila'))
        return context

class SessionCreateView(CreateView):
    model = ClassSession
    fields = ['session_date', 'start_time', 'end_time', 'topic', 'notes']

    def form_valid(self, form):
        class_instance = get_object_or_404(Class, id=self.kwargs['class_id'])
        form.instance.class_instance = class_instance
        form.instance.status = 'scheduled'
        self.object = form.save()

        # ✅ Return JSON if AJAX
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Session created successfully!',
                'session_id': str(self.object.id),
                'class_id': str(class_instance.id),
            })

        messages.success(self.request, 'Session created successfully!')
        return super().form_valid(form)

    def form_invalid(self, form):
        # ✅ Return JSON if AJAX
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'errors': form.errors,
            }, status=400)

        return super().form_invalid(form)

    def get(self, request, *args, **kwargs):
        # ✅ Prevent GET from accidentally returning HTML on AJAX call
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'error': 'GET not allowed for this endpoint'
            }, status=405)

        return super().get(request, *args, **kwargs)
@login_required
def start_session(request, session_id):
    import logging
    logger = logging.getLogger(__name__)

    session = get_object_or_404(ClassSession, id=session_id)
    logger.debug(f"Attempting to start session: {session.id}")
    logger.debug(f"Session status: {session.status}")
    logger.debug(f"Is past end time: {session.is_past_end_time()}")
    logger.debug(f"Session end datetime: {session.get_end_datetime()}")
    logger.debug(f"Now: {timezone.now()}")

    if session.status != 'scheduled':
        logger.warning(f"Session {session.id} not in 'scheduled' status. Current status: {session.status}")
        messages.error(request, 'Only scheduled sessions can be started')
    elif session.is_past_end_time():
        logger.warning(f"Session {session.id} end time has already passed.")
        messages.error(request, 'Cannot start session - end time has already passed')
    else:
        session.status = 'in_progress'
        session.save()
        logger.info(f"Session {session.id} started successfully.")
        messages.success(request, 'Session started successfully!')
    return redirect('session_list', class_id=session.class_instance.id)

@login_required
def complete_session(request, session_id):
    session = get_object_or_404(ClassSession, id=session_id)
    if session.status != 'in_progress':
        messages.error(request, 'Only in-progress sessions can be completed')
    else:
        session.status = 'completed'
        session.save()
        messages.success(request, 'Session completed successfully!')
    return redirect('session_list', class_id=session.class_instance.id)
@login_required
def take_attendance(request, session_id):
    """Handle attendance and participation marking for a session"""
    session = get_object_or_404(ClassSession, id=session_id)
    
    # Verify teacher owns this class
    if session.class_instance.teacher != request.user:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            with transaction.atomic():
                # Process Attendance
                if 'attendance' in data:
                    for student_id, status in data['attendance'].items():
                        if status in dict(Attendance.Status.choices).keys():
                            Attendance.objects.update_or_create(
                                session=session,
                                student_id=student_id,
                                defaults={
                                    'status': status,
                                    'marked_by': request.user,
                                    'marked_at': timezone.now()
                                }
                            )
                
                # Process Participation
                if 'participation' in data:
                    # Get or create default participation categories for this class
                    discussion_cat, _ = ParticipationCategory.objects.get_or_create(
                        class_instance=session.class_instance,
                        name='Discussion',
                        defaults={
                            'description': 'Class discussion participation',
                            'default_points': 1,
                            'color': '#3B82F6'
                        }
                    )
                    quiz_cat, _ = ParticipationCategory.objects.get_or_create(
                        class_instance=session.class_instance,
                        name='Quiz',
                        defaults={
                            'description': 'Quiz performance',
                            'default_points': 1,
                            'color': '#10B981'
                        }
                    )
                    activity_cat, _ = ParticipationCategory.objects.get_or_create(
                        class_instance=session.class_instance,
                        name='Activity',
                        defaults={
                            'description': 'Class activity participation',
                            'default_points': 1,
                            'color': '#F59E0B'
                        }
                    )
                    other_cat, _ = ParticipationCategory.objects.get_or_create(
                        class_instance=session.class_instance,
                        name='Other',
                        defaults={
                            'description': 'Other participation',
                            'default_points': 1,
                            'color': '#64748B'
                        }
                    )
                    
                    for student_id, categories in data['participation'].items():
                        # Save each category participation separately
                        for category_name, points in categories.items():
                            if points > 0:  # Only save if points awarded
                                category = {
                                    'discussion': discussion_cat,
                                    'quiz': quiz_cat,
                                    'activity': activity_cat,
                                    'other': other_cat
                                }.get(category_name.lower())
                                
                                if category:
                                    Participation.objects.update_or_create(
                                        session=session,
                                        student_id=student_id,
                                        category=category,
                                        defaults={
                                            'points': points,
                                            'awarded_by': request.user
                                        }
                                    )
            
            return JsonResponse({
                'success': True,
                'message': 'Data saved successfully'
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    # GET request - show combined page
    students = User.objects.filter(
        enrollment__class_instance=session.class_instance,
        enrollment__status='active',
        user_type='student'
    ).order_by('last_name', 'first_name')
    
    # Get existing records
    attendance_records = {
        record.student_id: record 
        for record in Attendance.objects.filter(session=session)
    }
    
    participation_records = Participation.objects.filter(
        session=session
    ).select_related('category')
    
    # Organize participation by student and category
    student_participation = defaultdict(dict)
    for record in participation_records:
        student_participation[record.student_id][record.category.name.lower()] = record.points
    
    students_data = []
    for student in students:
        attendance = attendance_records.get(student.id)
        participation = student_participation.get(student.id, {})
        
        students_data.append({
            'id': student.id,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'avatar_url': student.avatar_url or '/static/default-avatar.png',
            'attendance': {
                'status': attendance.status if attendance else None,
                'marked_at': attendance.marked_at if attendance else None
            },
            'participation': {
                'discussion': participation.get('discussion', 0),
                'quiz': participation.get('quiz', 0),
                'activity': participation.get('activity', 0),
                'other': participation.get('other', 0)
            }
        })
    
    context = {
        'session': session,
        'students': students_data,
        'class_obj': session.class_instance,
        'session_date': session.session_date.strftime("%b %d, %Y"),
        'session_time': f"{session.start_time.strftime('%I:%M %p').lstrip('0')} - {session.end_time.strftime('%I:%M %p').lstrip('0')}",
    }
    
    return render(request, 'classroom/teacher/take_attendance.html', context)

@login_required
def student_participation_report(request, class_id, student_id):
    # Get the class and verify teacher owns it
    class_obj = get_object_or_404(Class, id=class_id, teacher=request.user)
    
    # Get the student and verify they're enrolled
    student = get_object_or_404(User, id=student_id, user_type='student')
    enrollment = get_object_or_404(
        Enrollment, 
        class_instance=class_obj, 
        student=student,
        status='active'
    )
    
    # Get all participation records for this student in this class
    participations = Participation.objects.filter(
        student=student,
        session__class_instance=class_obj
    ).select_related('session', 'category').order_by('-session__session_date')
    
    # Get category totals
    categories = ParticipationCategory.objects.filter(
        class_instance=class_obj
    )
    
    category_totals = []
    for category in categories:
        total = participations.filter(category=category).aggregate(
            total=Sum('points')
        )['total'] or 0
        category_totals.append({
            'category': category,
            'total_points': total
        })
    
    # Calculate overall total
    total_points = sum(item['total_points'] for item in category_totals)
    
    context = {
        'class': class_obj,
        'student': student,
        'enrollment': enrollment,
        'participations': participations,
        'category_totals': category_totals,
        'total_points': total_points,
    }
    
    return render(request, 'classroom/teacher/student_participation_report.html', context)

@login_required
def student_attendance_report(request, class_id, student_id):
    # Get class and verify ownership
    class_obj = get_object_or_404(Class, id=class_id, teacher=request.user)
    student = get_object_or_404(User, id=student_id, user_type='student')
    
    # Get all completed sessions
    sessions = ClassSession.objects.filter(
        class_instance=class_obj,
        status='completed'
    ).order_by('-session_date')
    
    # Get attendance records in one query
    attendance_records = Attendance.objects.filter(
        student=student,
        session__in=sessions
    )
    
    # Calculate statistics
    present_count = attendance_records.filter(status='present').count()
    late_count = attendance_records.filter(status='late').count()
    absent_count = sessions.count() - present_count - late_count
    total_sessions = sessions.count()
    
    attendance_percentage = 0
    if total_sessions > 0:
        attendance_percentage = round((present_count / total_sessions) * 100, 1)
    
    # Prepare session data
    session_status = []
    for session in sessions:
        record = attendance_records.filter(session=session).first()
        session_status.append({
            'date': session.session_date,
            'topic': session.topic or "Class Session",
            'status': record.status if record else 'absent'
        })
    
    context = {
        'class': class_obj,
        'student': student,
        'student_name': f"{student.first_name} {student.last_name}",
        'class_name': class_obj.course_name,
        'session_status': session_status,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'total_sessions': total_sessions,
        'attendance_percentage': attendance_percentage,
    }
    return render(request, 'classroom/teacher/student_attendance_report.html', context)

@login_required
def generate_excel_report(request, class_id):
    class_obj = get_object_or_404(Class, id=class_id, teacher=request.user)
    report_type = request.GET.get('type', 'attendance')
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    # Filter data based on parameters
    sessions = ClassSession.objects.filter(
        class_instance=class_obj,
        status='completed'
    )
    
    if start_date and end_date:
        sessions = sessions.filter(
            session_date__range=[start_date, end_date]
        )
    
    # Create Excel workbook
    output = io.BytesIO()
    workbook = Workbook()
    
    # Attendance Sheet
    if report_type in ['attendance', 'combined']:
        attendance_sheet = workbook.active
        attendance_sheet.title = "Attendance"
        
        # Add headers
        headers = ["Student Name", "Present", "Absent", "Late", "Attendance %"]
        for col_num, header in enumerate(headers, 1):
            cell = attendance_sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Add data
        row_num = 2
        for enrollment in Enrollment.objects.filter(class_instance=class_obj, status='active'):
            student = enrollment.student
            attendance_records = Attendance.objects.filter(
                student=student,
                session__in=sessions
            )
            
            present = attendance_records.filter(status='present').count()
            absent = attendance_records.filter(status='absent').count()
            late = attendance_records.filter(status='late').count()
            total = present + absent + late
            percentage = (present / total * 100) if total > 0 else 0
            
            attendance_sheet.cell(row=row_num, column=1, value=f"{student.first_name} {student.last_name}")
            attendance_sheet.cell(row=row_num, column=2, value=present)
            attendance_sheet.cell(row=row_num, column=3, value=absent)
            attendance_sheet.cell(row=row_num, column=4, value=late)
            attendance_sheet.cell(row=row_num, column=5, value=percentage/100).number_format = '0.00%'
            row_num += 1
    
    # Participation Sheet
    if report_type in ['participation', 'combined']:
        participation_sheet = workbook.create_sheet("Participation")
        
        # Add headers
        headers = ["Student Name", "Discussion", "Quiz", "Activity", "Other", "Total Points"]
        for col_num, header in enumerate(headers, 1):
            cell = participation_sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Add data
        row_num = 2
        for enrollment in Enrollment.objects.filter(class_instance=class_obj, status='active'):
            student = enrollment.student
            participation = Participation.objects.filter(
                student=student,
                session__in=sessions
            )
            
            categories = {
                'discussion': 0,
                'quiz': 0,
                'activity': 0,
                'other': 0
            }
            
            for record in participation:
                cat_name = record.category.name.lower()
                if cat_name in categories:
                    categories[cat_name] += record.points
            
            participation_sheet.cell(row=row_num, column=1, value=f"{student.first_name} {student.last_name}")
            participation_sheet.cell(row=row_num, column=2, value=categories['discussion'])
            participation_sheet.cell(row=row_num, column=3, value=categories['quiz'])
            participation_sheet.cell(row=row_num, column=4, value=categories['activity'])
            participation_sheet.cell(row=row_num, column=5, value=categories['other'])
            participation_sheet.cell(row=row_num, column=6, value=sum(categories.values()))
            row_num += 1
    
    # Auto-adjust column widths
    for sheet in workbook:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"class_{class_id}_{report_type}_report.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

@login_required
def qr_attendance(request, session_id):
    session = get_object_or_404(ClassSession, id=session_id)
    # Set expiry to session end time
    expiry = session.get_end_datetime()
    now = timezone.now()

    # Generate or get QRCode object
    qr_obj, created = QRCode.objects.get_or_create(
        session=session,
        defaults={
            'code': str(session.id),  # You can use a more secure random code if needed
            'expires_at': expiry,
            'is_active': True,
        }
    )
    # If already exists but expired, regenerate
    if not created and qr_obj.expires_at < now:
        qr_obj.code = str(session.id)
        qr_obj.expires_at = expiry
        qr_obj.is_active = True
        qr_obj.scan_count = 0
        qr_obj.save()

    # The URL students will scan to mark attendance
    qr_url = request.build_absolute_uri(
        reverse('student_qr_attendance', args=[qr_obj.code])
    )

    # Generate QR code image
    img = qrcode.make(qr_url)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    image_stream = buf.getvalue()
    buf.close()

    # Render a template with the QR code
    return render(request, 'classroom/teacher/qr_attendance.html', {
        'session': session,
        'qr_code': image_stream,
        'qr_url': qr_url,
        'expiry': expiry,
    })

from django.contrib import messages

@login_required
def student_qr_attendance(request, code):
    qr_obj = get_object_or_404(QRCode, code=code, is_active=True)
    session = qr_obj.session
    now = timezone.now()
    if now > qr_obj.expires_at:
        messages.error(request, "QR code expired.")
        return redirect('student_dashboard')

    # Check if attendance already exists for this user and session
    attendance_exists = Attendance.objects.filter(
        session=session,
        student=request.user
    ).exists()

    if attendance_exists:
        messages.warning(request, "You have already marked your attendance for this session.")
        return redirect('student_dashboard')

    # Mark attendance for the student
    Attendance.objects.update_or_create(
        session=session,
        student=request.user,
        defaults={'status': Attendance.Status.PRESENT, 'marked_at': now}
    )
    qr_obj.scan_count += 1
    qr_obj.save()
    messages.success(request, "Attendance marked successfully!")
    return redirect('student_dashboard')
